import os
from flask import (Flask, render_template, redirect, url_for, request,
                   flash, session, send_file, jsonify, abort)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import date, datetime
import io

from models import (db, User, Client, ClientIdentifier, ImportBatch,
                    RawCharge, Invoice, InvoiceLine, InvoiceRun, CompanySettings)
from importer import parse_file, match_charges_to_clients
from billing import generate_invoices, generate_pdf

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'synthesis-billing-secret-change-me')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///billing.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

db.init_app(app)

# Jinja2 globals
@app.context_processor
def inject_globals():
    from datetime import date
    return {'today': date.today()}
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def get_settings():
    s = CompanySettings.query.first()
    if not s:
        s = CompanySettings(company_name='Synthesis IT')
        db.session.add(s)
        db.session.commit()
    return s


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.check_password(request.form['password']):
            login_user(user)
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    from models import IgnoredKey
    from sqlalchemy import func as sqlfunc

    total_clients = Client.query.filter_by(active=True).count()
    total_invoices = Invoice.query.count()
    recent_invoices = Invoice.query.order_by(Invoice.created_at.desc()).limit(8).all()

    latest_period = (db.session.query(ImportBatch.billing_period)
        .filter(ImportBatch.file_type.in_(['gamma_bb','gamma_ces','gamma_ipdc','gamma_wlr']))
        .order_by(ImportBatch.billing_period.desc()).first())
    current_period = latest_period[0] if latest_period else None

    period_revenue = 0.0
    period_invoices = 0
    if current_period:
        invs = Invoice.query.filter_by(billing_period=current_period).all()
        period_revenue = sum(i.total for i in invs)
        period_invoices = len(invs)

    ignored_keys = set(i.source_key for i in IgnoredKey.query.all())
    unmatched_count = (RawCharge.query
        .filter(RawCharge.matched == False, RawCharge.invoiced == False,
                RawCharge.archived == False,
                ~RawCharge.source_key.in_(ignored_keys) if ignored_keys else True)
        .distinct(RawCharge.source_key).count())

    draft_count = Invoice.query.filter_by(status='draft').count()

    dupes = (db.session.query(ImportBatch.filename, sqlfunc.count(ImportBatch.id).label('cnt'))
             .group_by(ImportBatch.filename)
             .having(sqlfunc.count(ImportBatch.id) > 1)
             .all())

    return render_template('dashboard.html',
        total_clients=total_clients,
        total_invoices=total_invoices,
        recent_invoices=recent_invoices,
        current_period=current_period,
        period_revenue=period_revenue,
        period_invoices=period_invoices,
        unmatched_count=unmatched_count,
        draft_count=draft_count,
        duplicate_batches=dupes)


# ── Clients ───────────────────────────────────────────────────────────────────

@app.route('/clients')
@login_required
def clients():
    q = request.args.get('q', '')
    query = Client.query.filter_by(active=True)
    if q:
        # Search name, account ref, and identifier values
        from sqlalchemy import exists
        query = query.filter(
            Client.name.ilike(f'%{q}%') |
            Client.account_ref.ilike(f'%{q}%') |
            Client.id.in_(
                db.session.query(ClientIdentifier.client_id)
                .filter(ClientIdentifier.id_value.ilike(f'%{q}%'))
            )
        )
    clients = query.order_by(Client.name).all()
    return render_template('clients/list.html', clients=clients, q=q)

@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def client_new():
    if request.method == 'POST':
        c = Client(
            name=request.form['name'],
            address_line1=request.form.get('address_line1'),
            address_line2=request.form.get('address_line2'),
            city=request.form.get('city'),
            postcode=request.form.get('postcode'),
            contact_name=request.form.get('contact_name'),
            contact_email=request.form.get('contact_email'),
            contact_phone=request.form.get('contact_phone'),
            account_ref=request.form.get('account_ref'),
            vat_number=request.form.get('vat_number'),
            markup_pct=float(request.form.get('markup_pct', 30.0)),
            billing_day=int(request.form.get('billing_day', 1)),
            notes=request.form.get('notes'),
        )
        db.session.add(c)
        db.session.commit()
        flash(f'Client "{c.name}" created.', 'success')
        return redirect(url_for('client_detail', id=c.id))
    return render_template('clients/form.html', client=None)

@app.route('/clients/<int:id>')
@login_required
def client_detail(id):
    from collections import defaultdict
    c = db.get_or_404(Client, id)
    invoices = Invoice.query.filter_by(client_id=id).order_by(Invoice.created_at.desc()).limit(20).all()

    all_charges = RawCharge.query.filter_by(client_id=id, invoiced=False).all()

    # Split into calls (itemised) and rentals (summarised)
    calls = []
    rental_summary = defaultdict(lambda: {'cost': 0.0, 'credit': 0.0, 'qty': 0, 'charge_type': '', 'period': ''})

    for ch in all_charges:
        if ch.charge_type == 'Call':
            calls.append(ch)
        else:
            key = (ch.billing_period, ch.charge_type, ch.product_name or 'Service')
            rental_summary[key]['cost'] += ch.cost_amount * ch.quantity
            rental_summary[key]['credit'] += ch.credit_amount
            rental_summary[key]['qty'] += ch.quantity
            rental_summary[key]['charge_type'] = ch.charge_type
            rental_summary[key]['period'] = ch.billing_period

    # Sort calls by date/time
    calls.sort(key=lambda x: (x.billing_period, x.call_date or '', str(x.id)))

    rentals = []
    for (period, ctype, product), vals in sorted(rental_summary.items()):
        net = vals['cost'] - vals['credit']
        rentals.append({
            'period': period,
            'charge_type': ctype,
            'product': product,
            'qty': vals['qty'],
            'cost': net,
            'sell': net * (1 + c.markup_pct / 100),
        })

    # Totals
    call_cost = sum(ch.cost_amount for ch in calls)
    rental_cost = sum(r['cost'] for r in rentals)
    call_sell = call_cost * (1 + c.markup_pct / 100)
    rental_sell = rental_cost * (1 + c.markup_pct / 100)

    def fmt_duration(secs):
        if not secs:
            return '—'
        m, s = divmod(int(secs), 60)
        return f"{m}m {s:02d}s"

    return render_template('clients/detail.html', client=c, invoices=invoices,
                           calls=calls, rentals=rentals,
                           call_cost=call_cost, rental_cost=rental_cost,
                           call_sell=call_sell, rental_sell=rental_sell,
                           fmt_duration=fmt_duration)

@app.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    c = db.get_or_404(Client, id)
    if request.method == 'POST':
        c.name = request.form['name']
        c.address_line1 = request.form.get('address_line1')
        c.address_line2 = request.form.get('address_line2')
        c.city = request.form.get('city')
        c.postcode = request.form.get('postcode')
        c.contact_name = request.form.get('contact_name')
        c.contact_email = request.form.get('contact_email')
        c.contact_phone = request.form.get('contact_phone')
        c.account_ref = request.form.get('account_ref')
        c.vat_number = request.form.get('vat_number')
        c.markup_pct = float(request.form.get('markup_pct', 30.0))
        c.billing_day = int(request.form.get('billing_day', 1))
        c.notes = request.form.get('notes')
        db.session.commit()
        flash('Client updated.', 'success')
        return redirect(url_for('client_detail', id=id))
    return render_template('clients/form.html', client=c)

@app.route('/clients/<int:id>/identifiers/add', methods=['POST'])
@login_required
def add_identifier(id):
    db.get_or_404(Client, id)
    id_type = request.form.get('id_type', 'gamma_cli')
    description = request.form.get('description', '')

    # Support bulk add — newline or comma separated values
    raw = request.form.get('id_value', '')
    # Split on newlines and commas, strip whitespace
    import re
    values = [v.strip() for v in re.split(r'[\n,]+', raw) if v.strip()]

    added = 0
    skipped = 0
    for val in values:
        existing = ClientIdentifier.query.filter_by(
            id_type=id_type, id_value=val).first()
        if existing:
            skipped += 1
        else:
            db.session.add(ClientIdentifier(
                client_id=id, id_type=id_type,
                id_value=val, description=description,
            ))
            added += 1

    if added:
        db.session.commit()
        flash(f'Added {added} identifier(s).{" "+str(skipped)+" already existed." if skipped else ""}', 'success')
    elif skipped:
        flash(f'All {skipped} identifier(s) already assigned.', 'warning')
    return redirect(url_for('client_detail', id=id))

@app.route('/identifiers/<int:id>/delete', methods=['POST'])
@login_required
def delete_identifier(id):
    ident = db.get_or_404(ClientIdentifier, id)
    client_id = ident.client_id
    db.session.delete(ident)
    db.session.commit()
    flash('Identifier removed.', 'success')
    return redirect(url_for('client_detail', id=client_id))


# ── Imports ───────────────────────────────────────────────────────────────────

@app.route('/imports')
@login_required
def imports():
    batches = ImportBatch.query.order_by(ImportBatch.imported_at.desc()).all()
    return render_template('imports/list.html', batches=batches)

@app.route('/imports/upload', methods=['GET', 'POST'])
@login_required
def import_upload():
    if request.method == 'POST':
        files = request.files.getlist('files')
        results = []
        for f in files:
            if not f.filename:
                continue
            try:
                content = f.read().decode('utf-8-sig', errors='replace')
                file_type, period, records = parse_file(f.filename, content)

                batch = ImportBatch(
                    filename=f.filename,
                    file_type=file_type,
                    billing_period=period,
                    imported_by=current_user.id,
                    record_count=len(records),
                )
                db.session.add(batch)
                db.session.flush()

                # Auto-archive any older uninvoiced batches of the same file type
                older_batches = (ImportBatch.query
                    .filter(ImportBatch.file_type == file_type,
                            ImportBatch.id != batch.id)
                    .all())
                archived_count = 0
                for old_batch in older_batches:
                    updated = (RawCharge.query
                        .filter_by(batch_id=old_batch.id, invoiced=False, archived=False)
                        .update({'archived': True}))
                    archived_count += updated
                if archived_count:
                    db.session.flush()

                charge_objs = []
                for r in records:
                    c = RawCharge(
                        batch_id=batch.id,
                        billing_period=period or r.get('billing_period', ''),
                        **{k: v for k, v in r.items()
                           if k in ('source_key','product_name','charge_type','call_date',
                                    'call_duration','destination','description',
                                    'cost_amount','credit_amount','quantity','site_name','raw_json')}
                    )
                    db.session.add(c)
                    charge_objs.append(c)

                db.session.flush()
                matched = match_charges_to_clients(charge_objs, db.session)
                batch.matched_count = matched
                db.session.commit()

                results.append({'file': f.filename, 'type': file_type,
                                 'records': len(records), 'matched': matched,
                                 'period': period})
            except Exception as e:
                db.session.rollback()
                results.append({'file': f.filename, 'error': str(e)})

        return render_template('imports/results.html', results=results)
    return render_template('imports/upload.html')

@app.route('/imports/<int:id>/delete', methods=['POST'])
@login_required
def import_delete(id):
    batch = db.get_or_404(ImportBatch, id)
    # Only allow deletion if no charges are invoiced
    invoiced = RawCharge.query.filter_by(batch_id=id, invoiced=True).count()
    if invoiced:
        flash('Cannot delete — some charges in this batch are already invoiced.', 'danger')
        return redirect(url_for('imports'))
    RawCharge.query.filter_by(batch_id=id).delete()
    db.session.delete(batch)
    db.session.commit()
    flash('Import batch deleted.', 'success')
    return redirect(url_for('imports'))


# ── Charges / Matching ────────────────────────────────────────────────────────

@app.route('/charges')
@login_required
def charges():
    from models import IgnoredKey
    from sqlalchemy import func
    period = request.args.get('period', '')
    client_id = request.args.get('client_id', '')
    unmatched_only = request.args.get('unmatched', '')

    # Get ignored keys
    ignored_keys = set(i.source_key for i in IgnoredKey.query.all())

    if unmatched_only:
        # Group unmatched by source key — one row per unknown identifier
        grouped = (db.session.query(
                RawCharge.source_key,
                func.count(RawCharge.id).label('count'),
                func.sum(RawCharge.cost_amount * RawCharge.quantity).label('total_cost'),
                func.min(ImportBatch.billing_period).label('period'),
                func.min(RawCharge.product_name).label('sample_product'),
                func.min(ImportBatch.file_type).label('file_type'),
            )
            .join(ImportBatch, RawCharge.batch_id == ImportBatch.id)
            .filter(RawCharge.matched == False,
                    RawCharge.invoiced == False,
                    RawCharge.archived == False,
                    ~RawCharge.source_key.in_(ignored_keys) if ignored_keys else True)
            .group_by(RawCharge.source_key)
            .order_by(func.count(RawCharge.id).desc())
            .all())

        clients = Client.query.filter_by(active=True).order_by(Client.name).all()
        periods = db.session.query(ImportBatch.billing_period).distinct().order_by(ImportBatch.billing_period.desc()).all()
        unmatched_count = len(grouped)

        return render_template('charges/unmatched.html',
            grouped=grouped, clients=clients,
            periods=[p[0] for p in periods if p[0]],
            unmatched_count=unmatched_count,
            ignored_keys=ignored_keys)

    query = RawCharge.query.join(ImportBatch, RawCharge.batch_id == ImportBatch.id)
    if period:
        query = query.filter(ImportBatch.billing_period == period)
    if client_id:
        query = query.filter(RawCharge.client_id == int(client_id))

    charges = query.order_by(RawCharge.id.desc()).limit(500).all()
    periods = db.session.query(ImportBatch.billing_period).distinct().order_by(ImportBatch.billing_period.desc()).all()
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()
    unmatched_count = (RawCharge.query
        .filter(RawCharge.matched == False, RawCharge.invoiced == False, RawCharge.archived == False,
                ~RawCharge.source_key.in_(ignored_keys) if ignored_keys else True)
        .distinct(RawCharge.source_key).count())

    return render_template('charges/list.html',
        charges=charges, periods=[p[0] for p in periods if p[0]],
        clients=clients, selected_period=period,
        selected_client=client_id, unmatched_only=unmatched_only,
        unmatched_count=unmatched_count)

@app.route('/charges/<int:id>/assign', methods=['POST'])
@login_required
def charge_assign(id):
    charge = db.get_or_404(RawCharge, id)
    client_id = request.form.get('client_id')
    if client_id:
        charge.client_id = int(client_id)
        charge.matched = True
        db.session.commit()
        flash('Charge assigned to client.', 'success')
    return redirect(request.referrer or url_for('charges'))

@app.route('/charges/rematch', methods=['POST'])
@login_required
def charge_rematch():
    """Re-run matching on all unmatched charges using current identifiers."""
    from models import IgnoredKey, ClientIdentifier

    # Force fresh read from DB
    db.session.expire_all()

    ignored_keys = set(i.source_key for i in IgnoredKey.query.all())

    # Build lookup fresh
    lookup = {}
    for ident in ClientIdentifier.query.filter_by(active=True).all():
        lookup[ident.id_value.upper().strip()] = ident.client_id

    unmatched = RawCharge.query.filter_by(matched=False, invoiced=False, archived=False).all()

    matched = 0
    for charge in unmatched:
        if charge.source_key in ignored_keys:
            continue
        key = (charge.source_key or '').upper().strip()
        if key in lookup:
            charge.client_id = lookup[key]
            charge.matched = True
            matched += 1

    db.session.commit()
    flash(f'Re-matched {matched} of {len(unmatched)} unmatched charges.', 'success')
    return redirect(url_for('charges', unmatched='1'))

@app.route('/charges/ignore-zero', methods=['POST'])
@login_required
def charge_ignore_zero():
    """Ignore all unmatched source keys that have zero total cost."""
    from models import IgnoredKey
    from sqlalchemy import func
    zero_keys = (db.session.query(RawCharge.source_key)
                 .filter(RawCharge.matched == False, RawCharge.invoiced == False, RawCharge.archived == False)
                 .group_by(RawCharge.source_key)
                 .having(func.sum(RawCharge.cost_amount) == 0)
                 .all())
    added = 0
    for (key,) in zero_keys:
        if key and not IgnoredKey.query.filter_by(source_key=key).first():
            db.session.add(IgnoredKey(source_key=key, ignored_by=current_user.id,
                                      reason='Auto-ignored: zero cost'))
            added += 1
    db.session.commit()
    flash(f'Ignored {added} zero-cost source key(s).', 'success')
    return redirect(url_for('charges', unmatched='1'))

@app.route('/charges/ignore-selected', methods=['POST'])
@login_required
def charge_ignore_selected():
    """Ignore a comma-separated list of source keys."""
    from models import IgnoredKey
    keys = request.form.get('keys', '').split(',')
    added = 0
    for key in keys:
        key = key.strip()
        if key and not IgnoredKey.query.filter_by(source_key=key).first():
            db.session.add(IgnoredKey(source_key=key, ignored_by=current_user.id,
                                      reason='Bulk ignored'))
            added += 1
    db.session.commit()
    flash(f'Ignored {added} source key(s).', 'success')
    return redirect(url_for('charges', unmatched='1'))

@app.route('/charges/ignore', methods=['POST'])
@login_required
def charge_ignore():
    from models import IgnoredKey
    source_key = request.form.get('source_key')
    if source_key:
        existing = IgnoredKey.query.filter_by(source_key=source_key).first()
        if not existing:
            db.session.add(IgnoredKey(
                source_key=source_key,
                reason=request.form.get('reason', ''),
                ignored_by=current_user.id
            ))
            db.session.commit()
        flash(f'"{source_key}" will no longer appear in unmatched charges.', 'success')
    return redirect(url_for('charges', unmatched='1'))

@app.route('/charges/unignore', methods=['POST'])
@login_required
def charge_unignore():
    from models import IgnoredKey
    source_key = request.form.get('source_key')
    IgnoredKey.query.filter_by(source_key=source_key).delete()
    db.session.commit()
    flash(f'"{source_key}" restored to unmatched charges.', 'success')
    return redirect(url_for('charges', unmatched='1'))

@app.route('/charges/ignored')
@login_required
def charges_ignored():
    from models import IgnoredKey
    ignored = IgnoredKey.query.order_by(IgnoredKey.ignored_at.desc()).all()
    return render_template('charges/ignored.html', ignored=ignored)

@app.route('/charges/bulk-assign', methods=['POST'])
@login_required
def charge_bulk_assign():
    source_key = request.form.get('source_key')
    client_id = int(request.form.get('client_id'))
    create_ident = request.form.get('create_identifier') == '1'
    id_type = request.form.get('id_type', 'gamma_circuit')

    updated = (RawCharge.query
               .filter_by(source_key=source_key, matched=False)
               .update({'client_id': client_id, 'matched': True}))

    if create_ident and source_key:
        existing = ClientIdentifier.query.filter_by(id_type=id_type, id_value=source_key).first()
        if not existing:
            db.session.add(ClientIdentifier(
                client_id=client_id, id_type=id_type,
                id_value=source_key,
                description='Auto-created on bulk assign'))
    db.session.commit()
    flash(f'Assigned {updated} charges to client.', 'success')
    return redirect(url_for('charges', unmatched='1'))



# ── Recurring Charges ─────────────────────────────────────────────────────────

@app.route('/clients/<int:id>/recurring/add', methods=['POST'])
@login_required
def add_recurring(id):
    from models import RecurringCharge
    db.get_or_404(Client, id)
    rc = RecurringCharge(
        client_id=id,
        description=request.form.get('description', ''),
        category=request.form.get('category', 'Other'),
        unit_price=float(request.form.get('unit_price', 0)),
        unit_cost=float(request.form.get('unit_cost', 0)),
        vat_rate=float(request.form.get('vat_rate', 20.0)),
    )
    db.session.add(rc)
    db.session.commit()
    flash('Recurring charge added.', 'success')
    return redirect(url_for('client_detail', id=id))

@app.route('/recurring/<int:id>/delete', methods=['POST'])
@login_required
def delete_recurring(id):
    from models import RecurringCharge
    rc = db.get_or_404(RecurringCharge, id)
    client_id = rc.client_id
    db.session.delete(rc)
    db.session.commit()
    flash('Recurring charge removed.', 'success')
    return redirect(url_for('client_detail', id=client_id))

# ── Invoices ──────────────────────────────────────────────────────────────────

@app.route('/invoices')
@login_required
def invoices():
    status = request.args.get('status', '')
    period = request.args.get('period', '')
    q = Invoice.query
    if status: q = q.filter_by(status=status)
    if period: q = q.filter_by(billing_period=period)
    invs = q.order_by(Invoice.created_at.desc()).all()
    periods = db.session.query(Invoice.billing_period).distinct().order_by(Invoice.billing_period.desc()).all()
    return render_template('invoices/list.html', invoices=invs,
        periods=[p[0] for p in periods if p[0]],
        selected_period=period, selected_status=status)

@app.route('/invoices/billing-run', methods=['POST'])
@login_required
def billing_run():
    """One-click billing run — generates invoices for all clients using most recent period."""
    s = get_settings()

    # Find most recent billing period from imported service files
    latest = (db.session.query(ImportBatch.billing_period)
              .filter(ImportBatch.file_type.in_(['gamma_bb','gamma_ces','gamma_ipdc','gamma_wlr','gamma_inb']))
              .order_by(ImportBatch.billing_period.desc())
              .first())

    if not latest:
        flash('No service files imported yet — import files first.', 'warning')
        return redirect(url_for('invoices'))

    period = latest[0]

    # Check not already run for this period
    existing = Invoice.query.filter_by(billing_period=period).first()
    if existing:
        flash(f'Invoices already exist for {period}. Delete them first to rerun.', 'warning')
        return redirect(url_for('invoices', period=period))

    client_ids = [c.id for c in Client.query.filter_by(active=True).all()]
    run = InvoiceRun(billing_period=period, created_by=current_user.id)
    db.session.add(run)
    db.session.flush()

    created = generate_invoices(period, client_ids, db.session, run.id, current_user.id, s)
    db.session.commit()
    flash(f'Billing run complete — {len(created)} invoices generated for {period}.', 'success')
    return redirect(url_for('invoices', period=period))

@app.route('/invoices/generate', methods=['GET', 'POST'])
@login_required
def invoice_generate():
    s = get_settings()
    if request.method == 'POST':
        period = request.form['billing_period']
        client_ids = request.form.getlist('client_ids')
        if not client_ids:
            # all clients with unmatched charges in period
            from sqlalchemy import distinct
                # All active clients get an invoice
            client_ids = [c.id for c in Client.query.filter_by(active=True).all()]

        run = InvoiceRun(billing_period=period, created_by=current_user.id)
        db.session.add(run)
        db.session.flush()

        created = generate_invoices(period, [int(i) for i in client_ids],
                                    db.session, run.id, current_user.id, s)
        db.session.commit()
        flash(f'Generated {len(created)} invoice(s) for {period}.', 'success')
        return redirect(url_for('invoices', period=period))

    # GET — show form
    periods = (db.session.query(ImportBatch.billing_period)
               .distinct().order_by(ImportBatch.billing_period.desc()).all())
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()
    return render_template('invoices/generate.html',
        periods=[p[0] for p in periods if p[0]], clients=clients)

@app.route('/invoices/bulk-delete', methods=['POST'])
@login_required
def invoices_bulk_delete():
    ids = request.form.get('ids', '').split(',')
    deleted = 0
    for inv_id in ids:
        try:
            inv = db.session.get(Invoice, int(inv_id.strip()))
            if inv and inv.status != 'paid':
                for line in inv.lines:
                    RawCharge.query.filter_by(invoice_line_id=line.id).update(
                        {'invoiced': False, 'invoice_line_id': None})
                db.session.delete(inv)
                deleted += 1
        except (ValueError, TypeError):
            continue
    db.session.commit()
    flash(f'Deleted {deleted} invoice(s).', 'success')
    return redirect(url_for('invoices'))

@app.route('/invoices/<int:id>')
@login_required
def invoice_detail(id):
    inv = db.get_or_404(Invoice, id)
    return render_template('invoices/detail.html', invoice=inv, settings=get_settings())

@app.route('/invoices/mark-all-sent', methods=['POST'])
@login_required
def invoices_mark_all_sent():
    period = request.form.get('period', '')
    q = Invoice.query.filter_by(status='draft')
    if period:
        q = q.filter_by(billing_period=period)
    updated = 0
    for inv in q.all():
        inv.status = 'sent'
        updated += 1
    db.session.commit()
    flash(f'Marked {updated} invoice(s) as sent.', 'success')
    return redirect(url_for('invoices', period=period))

@app.route('/invoices/<int:id>/mark-sent', methods=['POST'])
@login_required
def invoice_mark_sent(id):
    inv = db.get_or_404(Invoice, id)
    inv.status = 'sent'
    db.session.commit()
    flash(f'Invoice {inv.invoice_number} marked as sent.', 'success')
    return redirect(url_for('invoice_detail', id=id))

@app.route('/invoices/<int:id>/pdf')
@login_required
def invoice_pdf(id):
    inv = db.get_or_404(Invoice, id)
    s = get_settings()
    pdf_bytes = generate_pdf(inv, s)
    return send_file(
        io.BytesIO(bytes(pdf_bytes)),
        mimetype='application/pdf',
        as_attachment=request.args.get('download') == '1',
        download_name=f"Invoice_{inv.invoice_number}.pdf"
    )

@app.route('/invoices/<int:id>/status', methods=['POST'])
@login_required
def invoice_status(id):
    inv = db.get_or_404(Invoice, id)
    inv.status = request.form['status']
    db.session.commit()
    flash(f'Invoice marked as {inv.status}.', 'success')
    return redirect(url_for('invoice_detail', id=id))

@app.route('/invoices/<int:id>/delete', methods=['POST'])
@login_required
def invoice_delete(id):
    inv = db.get_or_404(Invoice, id)
    if inv.status == 'paid':
        flash('Cannot delete a paid invoice.', 'danger')
        return redirect(url_for('invoice_detail', id=id))
    # Unmark charges
    for line in inv.lines:
        RawCharge.query.filter_by(invoice_line_id=line.id).update(
            {'invoiced': False, 'invoice_line_id': None})
    db.session.delete(inv)
    db.session.commit()
    flash('Invoice deleted.', 'success')
    return redirect(url_for('invoices'))


# ── Settings ──────────────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    s = get_settings()
    users = User.query.all()
    if request.method == 'POST':
        for field in ('company_name','address_line1','address_line2','city','postcode',
                      'phone','email','website','vat_number','company_number',
                      'bank_name','bank_sort_code','bank_account','invoice_prefix'):
            setattr(s, field, request.form.get(field, ''))
        s.default_markup_pct = float(request.form.get('default_markup_pct', 30))
        s.default_vat_rate = float(request.form.get('default_vat_rate', 20))
        s.payment_terms_days = int(request.form.get('payment_terms_days', 30))
        db.session.commit()
        flash('Settings saved.', 'success')
        return redirect(url_for('settings'))
    return render_template('settings.html', s=s, users=users)

@app.route('/settings/users/new', methods=['POST'])
@login_required
def user_new():
    if current_user.role != 'admin':
        abort(403)
    u = User(username=request.form['username'], email=request.form['email'],
             role=request.form.get('role', 'user'))
    u.set_password(request.form['password'])
    db.session.add(u)
    db.session.commit()
    flash(f'User {u.username} created.', 'success')
    return redirect(url_for('settings'))

@app.route('/settings/users/<int:id>/delete', methods=['POST'])
@login_required
def user_delete(id):
    if current_user.role != 'admin':
        abort(403)
    if current_user.id == id:
        flash('Cannot delete your own account.', 'danger')
        return redirect(url_for('settings'))
    u = db.get_or_404(User, id)
    db.session.delete(u)
    db.session.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('settings'))



# ── Billing Summary ───────────────────────────────────────────────────────────

@app.route('/summary')
@login_required
def billing_summary():
    from collections import defaultdict
    from models import IgnoredKey

    period = request.args.get('period', '')
    periods = (db.session.query(ImportBatch.billing_period)
               .distinct().order_by(ImportBatch.billing_period.desc()).all())
    period_list = [p[0] for p in periods if p[0]]

    if not period and period_list:
        period = period_list[0]

    if not period:
        return render_template('summary.html', period=period, period_list=period_list,
                               rows=[], grand=None)

    ignored_keys = set(i.source_key for i in IgnoredKey.query.all())

    # Get all matched charges for this period
    charges = (RawCharge.query
               .join(ImportBatch, RawCharge.batch_id == ImportBatch.id)
               .filter(ImportBatch.billing_period == period,
                       RawCharge.matched == True,
                       RawCharge.client_id.isnot(None))
               .all())

    # Unmatched count for this period (excluding ignored)
    unmatched = (RawCharge.query
                 .join(ImportBatch, RawCharge.batch_id == ImportBatch.id)
                 .filter(ImportBatch.billing_period == period,
                         RawCharge.matched == False,
                         ~RawCharge.source_key.in_(ignored_keys) if ignored_keys else True)
                 .count())

    # Category mapping
    FILE_CAT = {
        'gamma_calls_sip': 'Calls', 'gamma_calls_div': 'Calls',
        'gamma_calls_ftc': 'Calls', 'gamma_calls_ibrs': 'Calls',
        'gamma_calls_nts': 'Calls', 'nasstar_cdr': 'Calls',
        'gamma_ipdc': 'SIP Trunks', 'gamma_bb': 'Broadband',
        'gamma_ces': 'Leased Lines', 'gamma_wlr': 'WLR',
        'gamma_inb': 'Inbound',
    }

    # Build per-client summary
    client_data = defaultdict(lambda: {
        'Calls': 0.0, 'Broadband': 0.0, 'Leased Lines': 0.0,
        'SIP Trunks': 0.0, 'WLR': 0.0, 'Inbound': 0.0, 'Other': 0.0,
        'total_cost': 0.0
    })

    batch_cache = {}
    for ch in charges:
        if ch.batch_id not in batch_cache:
            batch_cache[ch.batch_id] = db.session.get(ImportBatch, ch.batch_id)
        batch = batch_cache[ch.batch_id]
        cat = FILE_CAT.get(batch.file_type if batch else '', 'Other')
        net = ch.cost_amount * ch.quantity - ch.credit_amount
        client_data[ch.client_id][cat] += net
        client_data[ch.client_id]['total_cost'] += net

    # Build rows with client info
    rows = []
    cats = ['Calls', 'Broadband', 'Leased Lines', 'SIP Trunks', 'WLR', 'Inbound', 'Other']
    for client_id, data in sorted(client_data.items(),
                                   key=lambda x: (db.session.get(Client, x[0]).name if db.session.get(Client, x[0]) else '')):
        client = db.session.get(Client, client_id)
        if not client:
            continue
        markup = 1 + client.markup_pct / 100
        total_cost = data['total_cost']
        total_sell = total_cost * markup
        margin = total_sell - total_cost
        margin_pct = (margin / total_sell * 100) if total_sell else client.markup_pct
        rows.append({
            'client': client,
            'cats': {c: data[c] for c in cats},
            'total_cost': total_cost,
            'total_sell': total_sell,
            'margin': margin,
            'margin_pct': margin_pct,
        })

    # Grand totals
    grand = {
        'cats': {c: sum(r['cats'][c] for r in rows) for c in cats},
        'total_cost': sum(r['total_cost'] for r in rows),
        'total_sell': sum(r['total_sell'] for r in rows),
        'margin': sum(r['margin'] for r in rows),
    }
    grand['margin_pct'] = (grand['margin'] / grand['total_sell'] * 100) if grand['total_sell'] else 0

    return render_template('summary.html', period=period, period_list=period_list,
                           rows=rows, grand=grand, cats=cats, unmatched=unmatched)


@app.route('/summary/export-excel')
@login_required
def summary_export_excel():
    """Export billing summary as Excel for accounts."""
    from collections import defaultdict
    from models import IgnoredKey
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl not available', 'danger')
        return redirect(url_for('billing_summary'))

    period = request.args.get('period', '')
    if not period:
        return redirect(url_for('billing_summary'))

    ignored_keys = set(i.source_key for i in IgnoredKey.query.all())

    charges = (RawCharge.query
               .join(ImportBatch, RawCharge.batch_id == ImportBatch.id)
               .filter(ImportBatch.billing_period == period,
                       RawCharge.matched == True,
                       RawCharge.client_id.isnot(None))
               .all())

    FILE_CAT = {
        'gamma_calls_sip': 'Calls', 'gamma_calls_div': 'Calls',
        'gamma_calls_ftc': 'Calls', 'gamma_calls_ibrs': 'Calls',
        'gamma_calls_nts': 'Calls', 'nasstar_cdr': 'Calls',
        'gamma_ipdc': 'SIP Trunks', 'gamma_bb': 'Broadband',
        'gamma_ces': 'Leased Lines', 'gamma_wlr': 'WLR',
        'gamma_inb': 'Inbound',
    }

    client_data = defaultdict(lambda: {
        'Calls': 0.0, 'Broadband': 0.0, 'Leased Lines': 0.0,
        'SIP Trunks': 0.0, 'WLR': 0.0, 'Inbound': 0.0, 'Other': 0.0,
        'total_cost': 0.0
    })
    batch_cache = {}
    for ch in charges:
        if ch.batch_id not in batch_cache:
            batch_cache[ch.batch_id] = db.session.get(ImportBatch, ch.batch_id)
        batch = batch_cache[ch.batch_id]
        cat = FILE_CAT.get(batch.file_type if batch else '', 'Other')
        net = ch.cost_amount * ch.quantity - ch.credit_amount
        client_data[ch.client_id][cat] += net
        client_data[ch.client_id]['total_cost'] += net

    cats = ['Calls', 'Broadband', 'Leased Lines', 'SIP Trunks', 'WLR', 'Inbound', 'Other']

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f'Summary {period}'
    ws.sheet_view.showGridLines = False

    DARK = 'FF1E3A5F'
    MID  = 'FF2E6DA4'
    LGREY= 'FFF2F5F8'
    WHITE= 'FFFFFFFF'
    GREEN= 'FF059669'
    thin = Side(style='thin', color='FFE2E6EA')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, val, bg=DARK, fg='FFFFFFFF', bold=True, align='center'):
        cell.value = val
        cell.font = Font(name='Arial', bold=bold, color=fg, size=9)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = brd

    def data_cell(cell, val, fmt=None, bold=False, bg=WHITE, align='left'):
        cell.value = val
        cell.font = Font(name='Arial', bold=bold, size=9)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = brd
        if fmt:
            cell.number_format = fmt

    # Title
    ws.merge_cells('A1:N1')
    ws['A1'].value = f'Billing Summary — {period}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=DARK)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:N2')
    ws['A2'].value = f'Generated by Synthesis IT Billing System'
    ws['A2'].font = Font(name='Arial', size=8, color='FFAAAAAA')
    ws['A2'].fill = PatternFill('solid', fgColor=DARK)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 16

    # Headers row 3
    headers = ['Client', 'Account Ref'] + cats + ['Total Cost', 'Total Sell', 'Margin £', 'Margin %']
    col_widths = [28, 14] + [12]*7 + [12, 12, 12, 10]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr_cell(ws.cell(3, i), h, bg=MID)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30

    # Data rows
    rows_out = sorted(client_data.items(),
                      key=lambda x: (db.session.get(Client, x[0]).name if db.session.get(Client, x[0]) else ''))
    grand = {c: 0.0 for c in cats}
    grand.update({'total_cost': 0.0, 'total_sell': 0.0, 'margin': 0.0})

    for ri, (client_id, data) in enumerate(rows_out, 4):
        client = db.session.get(Client, client_id)
        if not client:
            continue
        bg = WHITE if ri % 2 == 0 else LGREY
        markup = 1 + client.markup_pct / 100
        total_cost = data['total_cost']
        total_sell = total_cost * markup
        margin = total_sell - total_cost
        margin_pct = (margin / total_sell * 100) if total_sell else client.markup_pct

        data_cell(ws.cell(ri, 1), client.name, bold=True, bg=bg)
        data_cell(ws.cell(ri, 2), client.account_ref or '', bg=bg, align='center')
        for ci, cat in enumerate(cats, 3):
            v = data[cat]
            data_cell(ws.cell(ri, ci), v if v else None, fmt='#,##0.0000', bg=bg, align='right')
        data_cell(ws.cell(ri, 10), total_cost, fmt='#,##0.00', bold=True, bg=bg, align='right')
        data_cell(ws.cell(ri, 11), total_sell, fmt='#,##0.00', bold=True, bg=bg, align='right')
        data_cell(ws.cell(ri, 12), margin, fmt='#,##0.00', bg=bg, align='right')
        data_cell(ws.cell(ri, 13), margin_pct/100, fmt='0.0%', bg=bg, align='right')

        for cat in cats:
            grand[cat] += data[cat]
        grand['total_cost'] += total_cost
        grand['total_sell'] += total_sell
        grand['margin'] += margin

    # Grand total row
    gr = len(rows_out) + 4
    ws.row_dimensions[gr].height = 18
    hdr_cell(ws.cell(gr, 1), 'GRAND TOTAL', bg=DARK, align='left')
    hdr_cell(ws.cell(gr, 2), '', bg=DARK)
    for ci, cat in enumerate(cats, 3):
        v = grand[cat]
        c = ws.cell(gr, ci)
        c.value = v if v else None
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', fgColor=DARK)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = brd
        c.number_format = '#,##0.0000'
    for ci, (val, fmt) in enumerate([(grand['total_cost'],'#,##0.00'),
                                      (grand['total_sell'],'#,##0.00'),
                                      (grand['margin'],'#,##0.00'),
                                      ((grand['margin']/grand['total_sell']*100/100) if grand['total_sell'] else 0,'0.0%')], 10):
        c = ws.cell(gr, ci)
        c.value = val
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', fgColor=DARK)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = brd
        c.number_format = fmt

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Billing_Summary_{period}.xlsx')


# ── API helpers ───────────────────────────────────────────────────────────────

@app.route('/api/unmatched-keys')
@login_required
def api_unmatched_keys():
    """Return unique source_keys that have unmatched charges, for the assign UI."""
    from sqlalchemy import func
    keys = (db.session.query(RawCharge.source_key,
                             func.count(RawCharge.id).label('count'),
                             func.sum(RawCharge.cost_amount).label('total'))
            .filter(RawCharge.matched == False, RawCharge.invoiced == False,
                    RawCharge.source_key.isnot(None))
            .group_by(RawCharge.source_key)
            .order_by(func.count(RawCharge.id).desc())
            .all())
    return jsonify([{'key': k, 'count': c, 'total': round(t or 0, 4)} for k, c, t in keys])


# ── Init ──────────────────────────────────────────────────────────────────────

def create_tables():
    with app.app_context():
        db.create_all()
        # Create default admin if no users exist
        if not User.query.first():
            admin = User(username='admin', email='admin@synthesisit.co.uk', role='admin')
            admin.set_password('changeme123')
            db.session.add(admin)
            # Default company settings
            s = CompanySettings(
                company_name='Synthesis IT',
                invoice_prefix='SIT',
                next_invoice_number=1001,
                default_markup_pct=30.0,
                default_vat_rate=20.0,
                payment_terms_days=30,
            )
            db.session.add(s)
            db.session.commit()
            print('Created default admin user: admin / changeme123')


if __name__ == '__main__':
    create_tables()
    app.run(debug=True, host='0.0.0.0', port=5000)
