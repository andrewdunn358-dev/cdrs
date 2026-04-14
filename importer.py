"""
Handles parsing of all supplier file formats:
  Gamma fixed charge files: BB, CES, IPDC, WLR, INB
  Gamma call CDR files:     SIP, DIV, FTC, IBRS, NTS  (_V3.txt)
  Nasstar CDR files:        .CDR
"""
import csv, json, re
from datetime import datetime, date
from io import StringIO

# ── file type detection ──────────────────────────────────────────────────────

def detect_file_type(filename, content_preview):
    """Return file_type string based on filename pattern."""
    fn = filename.upper()
    if fn.endswith('.CDR'):
        return 'nasstar_cdr'
    if 'BB.TXT' in fn or fn.endswith('BBFF.TXT') or '_BB' in fn:
        return 'gamma_bb'
    if 'CESFF' in fn or '_CES' in fn:
        return 'gamma_ces'
    if 'IPDCFF' in fn or '_IPDC' in fn:
        return 'gamma_ipdc'
    if 'WLRFF' in fn or '_WLR' in fn:
        return 'gamma_wlr'
    if 'INBFF' in fn or '_INB' in fn:
        return 'gamma_inb'
    if '_SIP_V3' in fn:
        return 'gamma_calls_sip'
    if '_DIV_V3' in fn:
        return 'gamma_calls_div'
    if '_FTC_V3' in fn:
        return 'gamma_calls_ftc'
    if '_IBRS_V3' in fn:
        return 'gamma_calls_ibrs'
    if '_NTS_V3' in fn:
        return 'gamma_calls_nts'
    return 'unknown'


# ── Gamma fixed charge files ─────────────────────────────────────────────────
# All share the same CSV format (no header row):
# BB/WLR/INB: connection_date, billing_period, circuit_id, product, charge_type, cost, credit, qty
# CES:        same + site_name at end
# IPDC:       same as BB

def _parse_gamma_ff(content, file_type):
    """Parse all Gamma FF flat-file formats into RawCharge dicts."""
    records = []
    reader = csv.reader(StringIO(content))
    for row in reader:
        if not row or not any(c.strip() for c in row):
            continue
        if len(row) < 7:
            continue
        try:
            cost = float(row[5]) if row[5] else 0.0
            credit = float(row[6]) if row[6] else 0.0
            qty = int(row[7]) if len(row) > 7 and row[7].strip().isdigit() else 1
            site_name = row[8].strip() if len(row) > 8 else ''

            charge_type = row[4].strip()
            if charge_type == 'Ceased':
                charge_type = 'Credit'

            # For IPDC files, Gamma puts the line total in the cost column
            # and the channel count in quantity — treat cost as total, qty as 1
            if file_type == 'gamma_ipdc':
                qty = 1

            records.append({
                'source_key': row[2].strip(),
                'product_name': row[3].strip(),
                'charge_type': charge_type,
                'billing_period': _parse_billing_period(row[1].strip()),
                'cost_amount': cost,
                'credit_amount': abs(credit),
                'quantity': qty,
                'site_name': site_name,
                'description': f"{row[3].strip()} — {row[2].strip()}",
                'raw_json': json.dumps(row),
            })
        except (ValueError, IndexError):
            continue
    return records


def _parse_billing_period(text):
    """Convert 'March 2026' or 'February 2026' to '2026-03'."""
    months = {'january':'01','february':'02','march':'03','april':'04',
              'may':'05','june':'06','july':'07','august':'08',
              'september':'09','october':'10','november':'11','december':'12'}
    parts = text.lower().split()
    if len(parts) == 2:
        m = months.get(parts[0])
        y = parts[1]
        if m and y:
            return f"{y}-{m}"
    return text


# ── Gamma call CDR files (_V3.txt) ───────────────────────────────────────────
# Header row present.
# Key cols: 2=customer_id, 3=destination, 4=call_date, 5=call_time,
#           6=duration, 10=chargecode, 11=timeband, 12=salesprice,
#           15=ddi, 26=diverted_number

def _parse_gamma_calls(content, file_type):
    records = []
    reader = csv.reader(StringIO(content))
    headers = None
    for row in reader:
        if headers is None:
            headers = row
            continue
        if not row or len(row) < 13:
            continue
        try:
            duration = int(float(row[6])) if row[6] else 0
            cost = float(row[12]) if row[12] else 0.0
            call_date = None
            try:
                call_date = datetime.strptime(row[4].strip(), '%d/%m/%Y').date()
            except Exception:
                pass

            # For SIP calls use the DDI (col 15) as source key so individual
            # clients on shared trunks match correctly via their phone numbers.
            # Fall back to endpoint ID (col 2) if DDI is empty.
            ddi_raw = row[15].strip() if len(row) > 15 else ''
            if ddi_raw.startswith('+44'):
                source_key = '0' + ddi_raw[3:]
            elif ddi_raw.startswith('+'):
                source_key = ddi_raw  # international
            elif ddi_raw:
                source_key = ddi_raw
            else:
                # No DDI — fall back to endpoint ID (col 2)
                source_key = row[2].strip()
                if source_key.startswith('+44'):
                    source_key = '0' + source_key[3:]
                elif source_key.startswith('+'):
                    pass
                elif source_key.startswith('44') and len(source_key) > 10:
                    source_key = '0' + source_key[2:]

            # normalise destination: +44xxxxxxx -> 0xxxxxxx
            dest_raw = row[3].strip()
            if dest_raw.startswith('+44'):
                dest = '0' + dest_raw[3:]
            elif dest_raw.startswith('+'):
                dest = dest_raw  # international, keep as-is
            else:
                dest = dest_raw

            # call time is in col 5
            call_time = row[5].strip() if len(row) > 5 else ''
            desc = row[9].strip() if len(row) > 9 else ''

            records.append({
                'source_key': source_key,
                'product_name': f"Call — {desc}" if desc else f"Call — {row[10].strip() if len(row)>10 else ''}",
                'charge_type': 'Call',
                'billing_period': '',
                'call_date': call_date,
                'call_duration': duration,
                'destination': dest,
                'description': call_time,  # store call time here for display
                'cost_amount': cost,
                'credit_amount': 0.0,
                'quantity': 1,
                'site_name': '',
                'raw_json': json.dumps(row),
            })
        except (ValueError, IndexError):
            continue
    return records


# ── Nasstar CDR files ────────────────────────────────────────────────────────
# Line 0: HEADER,{account_id}
# Data rows: switch_node, cli, datetime, destination, duration, dest_type, cost, cli_prefix, direction, rate_band, ...
# Last line: TRAILER,{count}
# Paired files: direction I = cost-bearing leg; direction O paired file has cost=0

def _parse_nasstar_cdr(content, filename):
    records = []
    lines = content.splitlines()
    if not lines:
        return records

    account_id = ''
    if lines[0].startswith('HEADER,'):
        account_id = lines[0].split(',')[1].strip()

    for line in lines[1:]:
        line = line.strip()
        if not line or line.startswith('TRAILER') or line.startswith('HEADER'):
            continue
        parts = line.split(',')
        if len(parts) < 9:
            continue
        try:
            cost = float(parts[6]) if parts[6] else 0.0
            duration = int(float(parts[4])) if parts[4] else 0
            direction = parts[8].strip()

            # Skip outbound legs of paired files (cost is always on the I leg)
            # But include if this is a standalone outbound-only file (no paired I file)
            # We detect this by checking if cost > 0 OR direction is O (single-leg format)
            if direction == 'I' and cost == 0.0:
                continue  # empty inbound legs

            call_dt = None
            try:
                call_dt = datetime.strptime(parts[2].strip(), '%d/%m/%y %H:%M:%S').date()
            except Exception:
                pass

            call_time = ''
            try:
                call_time = datetime.strptime(parts[2].strip(), '%d/%m/%y %H:%M:%S').strftime('%H:%M:%S')
            except Exception:
                pass

            dest_raw = parts[3].strip() if len(parts) > 3 else ''
            dest = dest_raw  # Nasstar destinations are already in 07xxx / 01xxx format

            dest_type = parts[5].strip() if len(parts) > 5 else ''
            desc_map = {
                'UKGEO': 'UK Geographic', 'UKN': 'UK National',
                'UKNFM1': 'UK Mobile (O2)', 'UKNFM3': 'UK Mobile (T-Mobile)',
                'UKNFM5': 'UK Mobile (Vodafone)', 'UKNFM6': 'UK Mobile (3)',
                '03UK': 'UK 03 Number', 'NAT': 'National',
            }
            description = desc_map.get(dest_type, dest_type)

            records.append({
                'source_key': account_id,
                'product_name': f"Call — {description}",
                'charge_type': 'Call',
                'billing_period': '',
                'call_date': call_dt,
                'call_duration': duration,
                'destination': dest,
                'description': call_time,
                'cost_amount': cost,
                'credit_amount': 0.0,
                'quantity': 1,
                'site_name': '',
                'raw_json': json.dumps(parts),
            })
        except (ValueError, IndexError):
            continue
    return records


# ── Public entry point ────────────────────────────────────────────────────────

def parse_file(filename, content):
    """
    Parse a supplier billing file.
    Returns (file_type, billing_period, list_of_charge_dicts).
    """
    file_type = detect_file_type(filename, content[:500])

    if file_type in ('gamma_bb', 'gamma_ces', 'gamma_ipdc', 'gamma_wlr', 'gamma_inb'):
        records = _parse_gamma_ff(content, file_type)
        period = records[0]['billing_period'] if records else ''
        return file_type, period, records

    elif file_type in ('gamma_calls_sip', 'gamma_calls_div', 'gamma_calls_ftc',
                       'gamma_calls_ibrs', 'gamma_calls_nts'):
        records = _parse_gamma_calls(content, file_type)
        # derive period from call dates
        dates = [r['call_date'] for r in records if r.get('call_date')]
        period = f"{dates[0].year}-{dates[0].month:02d}" if dates else ''
        return file_type, period, records

    elif file_type == 'nasstar_cdr':
        records = _parse_nasstar_cdr(content, filename)
        dates = [r['call_date'] for r in records if r.get('call_date')]
        period = f"{dates[0].year}-{dates[0].month:02d}" if dates else ''
        return file_type, period, records

    return file_type, '', []


def match_charges_to_clients(charges, session):
    """
    Given a list of RawCharge ORM objects, look up ClientIdentifier
    and set client_id + matched flag. Returns count matched.
    """
    from models import ClientIdentifier
    # Build lookup dict: (id_type_group, normalised_value) -> client_id
    lookup = {}
    for ident in session.query(ClientIdentifier).filter_by(active=True).all():
        lookup[ident.id_value.upper().strip()] = ident.client_id

    matched = 0
    for charge in charges:
        key = (charge.source_key or '').upper().strip()
        if key in lookup:
            charge.client_id = lookup[key]
            charge.matched = True
            matched += 1
    return matched


def _pl_rows(path_or_bytes, sheet_idx=0):
    """Load worksheet rows from file path or bytes."""
    import openpyxl, io
    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(path_or_bytes, read_only=True, data_only=True)
    return list(wb.worksheets[sheet_idx].iter_rows(values_only=True)), wb.sheetnames

def _find_header(rows, keyword='service'):
    for i, row in enumerate(rows[:20]):
        if row[1] and str(row[1]).lower().strip() == keyword:
            return i
    return None

def _safe_float(v):
    if v is None: return 0.0
    try: return float(str(v).replace('£','').replace(',','').strip())
    except: return 0.0

def _s(v): return str(v).strip() if v else ''

def parse_price_list(filename, content_bytes, list_type):
    """Parse any Gamma price list Excel. Returns list of entry dicts."""
    entries = []

    def entry(service, billing_name, charge_type, price, install=0, cease_in=0, cease_out=0, notes=''):
        return dict(service=service, billing_name=billing_name, charge_type=charge_type,
                    unit_price=_safe_float(price), install_price=_safe_float(install),
                    cease_price_in=_safe_float(cease_in), cease_price_out=_safe_float(cease_out),
                    notes=_s(notes))

    if list_type == 'broadband':
        rows, _ = _pl_rows(content_bytes)
        hi = _find_header(rows)
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or str(row[1]).isupper(): continue  # skip section headers
            try:
                svc = _s(row[1])
                billing = svc.split(' (')[0].strip()
                entries.append(entry(svc, billing, 'Rental',
                    row[6], row[5], row[9], row[10], row[12] if len(row)>12 else ''))
            except: continue

    elif list_type in ('sip', 'inbound', 'webex', 'horizon', 'gamma_plus'):
        rows, _ = _pl_rows(content_bytes)
        hi = _find_header(rows)
        if hi is None: return entries
        # Detect column layout
        headers = [_s(v).lower() for v in rows[hi]]
        price_col = next((i for i,h in enumerate(headers) if 'price' in h), 4)
        bname_col = 2
        ctype_col = 3
        for row in rows[hi+1:]:
            if not row[1] or not row[price_col]: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[bname_col]), _s(row[ctype_col]),
                    row[price_col], notes=_s(row[5] if len(row)>5 else '')))
            except: continue

    elif list_type == 'ethernet':
        rows, _ = _pl_rows(content_bytes)
        hi = _find_header(rows)
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[4] is None: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[2]), _s(row[3]),
                    row[4], notes=_s(row[5] if len(row)>5 else '')))
            except: continue

    elif list_type == 'wlr':
        # WLR has multiple sheets and different column layout
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            hi = None
            for i, row in enumerate(rows[:15]):
                if row[1] and 'connection' in _s(row[1]).lower():
                    hi = i
                    break
                if row[1] and _s(row[1]).lower() in ('service', 'new connection - basic single line'):
                    hi = i
                    break
            if hi is None: continue
            for row in rows[hi:]:
                if not row[1] or not row[6]: continue
                try:
                    price = _safe_float(row[6])
                    if price > 0:
                        entries.append(entry(_s(row[1]), _s(row[3]), _s(row[2]),
                            price, notes=_s(row[5] if len(row)>5 else '')))
                except: continue

    elif list_type == 'porting':
        # Porting has service in col1 OR col2, price in col5
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            hi = _find_header(rows)
            if hi is None: continue
            current_service = ""
            for row in rows[hi+1:]:
                if row[1]: current_service = _s(row[1])
                svc = current_service
                bname = _s(row[2]) if row[2] else svc
                if not bname or row[5] is None: continue
                try:
                    entries.append(entry(svc, bname, _s(row[4]), row[5],
                        notes=_s(row[6] if len(row)>6 else '')))
                except: continue

    elif list_type == 'mobile':
        rows, _ = _pl_rows(content_bytes)
        hi = None
        for i, row in enumerate(rows[:15]):
            if row[1] and _s(row[1]).lower() == 'service':
                hi = i; break
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[5] is None: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[2]), 'Rental',
                    row[5], notes=_s(row[4] if len(row)>4 else '')))
            except: continue

    elif list_type == 'phoneline':
        rows, _ = _pl_rows(content_bytes)
        hi = None
        for i, row in enumerate(rows[:15]):
            if row[1] and _s(row[1]).lower() in ('service', 'services'):
                hi = i; break
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[5] is None: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[2]), _s(row[4]),
                    row[5], notes=_s(row[6] if len(row)>6 else '')))
            except: continue

    elif list_type == 'admin':
        rows, _ = _pl_rows(content_bytes)
        hi = _find_header(rows)
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[3] is None: continue
            try:
                price = _safe_float(row[3])
                if price > 0:
                    entries.append(entry(_s(row[1]), _s(row[2]), 'Charge', price))
            except: continue

    elif list_type == 'intl_sip':
        rows, _ = _pl_rows(content_bytes)
        # Header has Service in col1, BillingName in col2, Price in col3, ChargeType in col4
        hi = None
        for i, row in enumerate(rows[:15]):
            if row[1] and 'sip' in _s(row[1]).lower():
                hi = i; break
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[3] is None: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[2]), _s(row[4] if len(row)>4 else 'Rental'), row[3]))
            except: continue

    elif list_type == 'safeweb':
        rows, _ = _pl_rows(content_bytes)
        # Header: Products | Term | Charge Type | Price
        hi = None
        for i, row in enumerate(rows[:15]):
            if row[1] and _s(row[1]).lower() in ('products', 'service', 'services'):
                hi = i; break
        if hi is None: return entries
        for row in rows[hi+1:]:
            if not row[1] or row[4] is None: continue
            try:
                entries.append(entry(_s(row[1]), _s(row[1]), _s(row[3]), row[4],
                    notes=_s(row[2] if len(row)>2 else '')))
            except: continue

    return [e for e in entries if e['unit_price'] >= 0]
